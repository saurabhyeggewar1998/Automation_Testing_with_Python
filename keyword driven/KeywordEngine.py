import driver as driver
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
# -------     Read Data from excel   ---------
from selenium.webdriver.common.by import By

import xutils

serv_obj = Service("F:/Selenium/chromedriver.exe")
wb = openpyxl.load_workbook('C:/Users/Dell/BinarySearch/HubSpot.xlsx')
sheets = wb.sheetnames

sheet = wb['login']
rows = sheet.max_row
cols = sheet.max_column

c = 1

for r in range(1, 9):
    # print(rows)
    print(r)
    locator_type = sheet.cell(r + 1, c + 1).value

    locator_value = sheet.cell(r + 1, c + 2).value
    action = sheet.cell(r + 1, c + 3).value
    value = sheet.cell(r + 1, c + 4).value
    print(value)
    # match action:
    #     case "open browser":
    #         driver = webdriver.Chrome(service=serv_obj)
    #         break
    #     case "enterUrl":
    #         print("Enter Url")
    #         print(action)
    #         driver.get(value)
    #         break
    #     case "quit":
    #         driver.quit()
    #         break
    #     case _:
    #         print("Defual Action")
    #         print(action)
    #
    # match locator_type:
    #     case "id":
    #         element = driver.find_element(By.ID(locator_value))
    #         if action == "sendKeys":
    #             element.clear()
    #             element.send_keys(value)
    #         elif action == "click":
    #             element.click()
    #         elif action == "isDisplyed":
    #             element.is_displayed()
    #         elif action == "getText":
    #             element.text()
    #         break
    #     case "name":
    #         element = driver.find_element(By.NAME(locator_value))
    #         if action == "sendKeys":
    #             element.clear()
    #             element.send_keys(value)
    #         elif action == "click":
    #             element.click()
    #         elif action == "isDisplyed":
    #             element.is_displayed()
    #         elif action == "getText":
    #             element.text()
    #         break
    #     case "xpath":
    #         element = driver.find_element(By.XPATH(locator_value))
    #         if action == "sendKeys":
    #             element.clear()
    #             element.send_keys(value)
    #         elif action == "click":
    #             element.click()
    #         elif action == "isDisplyed":
    #             element.is_displayed()
    #         elif action == "getText":
    #             element.text()
    #         break
    #     case "css":
    #         element = driver.find_element(By.CSS_SELECTOR(locator_value))
    #         if action == "sendKeys":
    #             element.clear()
    #             element.send_keys(value)
    #         elif action == "click":
    #             element.click()
    #         elif action == "isDisplyed":
    #             element.is_displayed()
    #         elif action == "getText":
    #             element.text()
    #         break
    #
    #     case "linkText":
    #         element = driver.find_element(By.LINK_TEXT(locator_value))
    #
    #         if action == "click":
    #             element.click()
    #         break
    #     case "partialLinkText":
    #         element = driver.find_element(By.PARTIAL_LINK_TEXT(locator_value))
    #
    #         if action == "click":
    #             element.click()
    #         break
    #     case _:
    #         break
print("hello Good mrng")
