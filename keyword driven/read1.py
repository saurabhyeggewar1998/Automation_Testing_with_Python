import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service


# import xutils1
# file = 'C:/Users/Dell/BinarySearch/Book3.xlsx'
# workbook = openpyxl.load_workbook(file)
# sheet = workbook['Practice']
#
# rows = sheet.max_row
# cols = sheet.max_column
# for r in range(2,rows+1):
#     for c in range(1, cols + 1):
#


wb = openpyxl.load_workbook('C:/Users/Dell/BinarySearch/Book3.xlsx')
# sheets = wb.sheetnames
# print(sheets)
sh1 = wb['Practice']
# data = sh1['A2'].value
# print(data)
n = 4
for i in range(0, n):
    print(i)
# option1
# print(wb['Practice']['A2'].value)
# print(sh1.cell(2, 2).value)
# print(sh1.cell(2, 3).value)
# print(sh1.cell(2, 4).value)
# print(sh1.cell(2, 5).value)
#
# print("================================")
# print(wb['Practice']['A3'].value)
# print(sh1.cell(3, 2).value)
# print(sh1.cell(3, 3).value)
# print(sh1.cell(3, 4).value)
# print(sh1.cell(3, 5).value)
#
# print("=====================================")
# print(wb['Practice']['A4'].value)
# print(sh1.cell(4, 2).value)
# print(sh1.cell(4, 3).value)
# print(sh1.cell(4, 4).value)
# print(sh1.cell(4, 5).value)
#
# print("==========================================")
# print(wb['Practice']['A5'].value)
# print(sh1.cell(5, 2).value)
# print(sh1.cell(5, 3).value)
# print(sh1.cell(5, 4).value)
print(sh1.cell(5, 5).value)

# if action=="open browser":
#     driver = webdriver.Chrome(service=serv_obj)
# elif action=="enter url":
#     driver.get(value)
#
# elif action=="quit":
#     driver.quit()
# else:
#      print("Defual Action")
#      print(action)